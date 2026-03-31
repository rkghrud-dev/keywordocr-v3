"""
쿠팡 WING API 클라이언트 + 업로드 파이프라인
- 키: ~/Desktop/key/ 폴더에서 로드
- GUI 연동을 위한 콜백 기반 인터페이스
"""
from __future__ import annotations

import hmac
import hashlib
import json
import os
import re
import ssl
import time
import urllib.request
import urllib.parse
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Callable

# ─── 키/설정 로드 ───────────────────────────────
_KEY_DIR = os.path.join(os.path.expanduser("~"), "Desktop", "key")

def _load_key(filename: str, fallback: str = "") -> str:
    path = os.path.join(_KEY_DIR, filename)
    if os.path.isfile(path):
        return open(path, encoding="utf-8").read().strip()
    return fallback

def _load_kv_file(filename: str) -> dict[str, str]:
    """KEY=VALUE 형식 파일 로드"""
    path = os.path.join(_KEY_DIR, filename)
    kv = {}
    if os.path.isfile(path):
        for line in open(path, encoding="utf-8"):
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                kv[k.strip()] = v.strip()
    return kv

_COUPANG_KEYS = _load_kv_file("coupang_wing_api.txt")
ACCESS_KEY = _COUPANG_KEYS.get("access_key", "e747c511-5917-46f6-a601-209db9719d5e")
SECRET_KEY = _COUPANG_KEYS.get("secret_key", "ed887699fd65bad93f1195513ae3bdb12bc94be7")
VENDOR_ID = _COUPANG_KEYS.get("vendor_id", "A00704210")

OUTBOUND_CODE = 23273329
RETURN_CENTER_CODE = 1002256451


# ─── API 클라이언트 ─────────────────────────────

def _auth(method: str, path: str, query: str | None = None) -> str:
    now = datetime.now(timezone.utc)
    dt = now.strftime("%y%m%dT%H%M%SZ")
    message = dt + method + path + (query or "")
    sig = hmac.new(SECRET_KEY.encode(), message.encode(), hashlib.sha256).hexdigest()
    return f"CEA algorithm=HmacSHA256, access-key={ACCESS_KEY}, signed-date={dt}, signature={sig}"


def api_call(method: str, path: str, query: str | None = None, body: bytes | None = None) -> dict:
    import gzip as _gzip
    url = f"https://api-gateway.coupang.com{path}"
    if query:
        url += f"?{query}"
    req = urllib.request.Request(url)
    req.add_header("Content-type", "application/json;charset=UTF-8")
    req.add_header("Authorization", _auth(method, path, query))
    req.add_header("X-EXTENDED-TIMEOUT", "90000")
    req.add_header("Accept-Encoding", "gzip, identity")
    req.get_method = lambda: method

    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE

    def _decode_resp(resp_obj) -> dict:
        raw = resp_obj.read()
        if raw[:2] == b"\x1f\x8b":
            raw = _gzip.decompress(raw)
        return json.loads(raw.decode("utf-8"))

    try:
        resp = urllib.request.urlopen(req, body, context=ctx, timeout=30) if body else urllib.request.urlopen(req, context=ctx, timeout=30)
        return _decode_resp(resp)
    except urllib.request.HTTPError as e:
        raw = e.read()
        if raw[:2] == b"\x1f\x8b":
            try:
                raw = _gzip.decompress(raw)
            except Exception:
                pass
        msg = raw.decode("utf-8", errors="replace")
        return {"_error": e.code, "_msg": msg[:500]}


def predict_category(product_name: str) -> dict:
    body = json.dumps({"productName": product_name}).encode("utf-8")
    return api_call("POST", "/v2/providers/openapi/apis/api/v1/categorization/predict", body=body)


def get_category_meta(category_code: int) -> dict:
    path = f"/v2/providers/seller_api/apis/api/v1/marketplace/meta/category-related-metas/display-category-codes/{category_code}"
    return api_call("GET", path)


def create_product(product_json: dict) -> dict:
    body = json.dumps(product_json).encode("utf-8")
    return api_call("POST", "/v2/providers/seller_api/apis/api/v1/marketplace/seller-products", body=body)


# ─── 엑셀 → 쿠팡 JSON 변환 ────────────────────

_IMAGE_SELECTION_CACHE: dict[str, dict[str, dict[str, object]]] = {}
_COUPANG_PUBLIC_IMAGE_CACHE: dict[str, str] = {}


def _extract_gs9(code: str | None) -> str:
    match = re.search(r"GS\d{7,9}", str(code or ""), re.IGNORECASE)
    return match.group(0).upper() if match else ""


def _resolve_export_root(source_file_path: str) -> str:
    from pathlib import Path

    path = Path(source_file_path).resolve()
    parent = path.parent
    parent_name = parent.name.lower()
    grand_name = parent.parent.name.lower() if parent.parent else ""

    if parent_name == "llm_result" and grand_name == "llm_chunks":
        return str(parent.parent.parent)
    if parent_name == "llm_result":
        return str(parent.parent)
    return str(parent)


def _load_image_selections(export_root: str) -> dict[str, dict[str, object]]:
    cached = _IMAGE_SELECTION_CACHE.get(export_root)
    if cached is not None:
        return cached

    path = os.path.join(export_root, "image_selections.json")
    if not os.path.isfile(path):
        _IMAGE_SELECTION_CACHE[export_root] = {}
        return _IMAGE_SELECTION_CACHE[export_root]

    try:
        with open(path, encoding="utf-8") as f:
            raw = json.load(f)
        parsed: dict[str, dict[str, object]] = {}
        for key, value in raw.items():
            if not isinstance(value, dict):
                continue
            parsed[key.upper()] = {
                "main": value.get("main"),
                "mainB": value.get("mainB"),
                "additional": value.get("additional") if isinstance(value.get("additional"), list) else [],
            }
        _IMAGE_SELECTION_CACHE[export_root] = parsed
    except Exception:
        _IMAGE_SELECTION_CACHE[export_root] = {}

    return _IMAGE_SELECTION_CACHE[export_root]


def _find_listing_folder(export_root: str, gs9: str) -> str | None:
    from pathlib import Path

    for root_name in ("listing_images",):
        base = Path(export_root) / root_name
        if not base.is_dir():
            continue
        matches = sorted((p for p in base.rglob(gs9) if p.is_dir()), key=lambda p: str(p))
        if matches:
            return str(matches[0])
    return None


def _pick_local_listing_images(row: dict) -> list[str]:
    source_file_path = str(row.get("_source_file_path") or "")
    gs9 = _extract_gs9(row.get("자체 상품코드"))
    if not source_file_path or not gs9:
        return []

    export_root = str(row.get("_export_root") or _resolve_export_root(source_file_path))
    folder_path = _find_listing_folder(export_root, gs9)
    if not folder_path:
        return []

    files = sorted(
        [
            os.path.join(folder_path, name)
            for name in os.listdir(folder_path)
            if os.path.splitext(name)[1].lower() in (".jpg", ".jpeg", ".png", ".bmp", ".webp")
        ],
        key=lambda p: os.path.basename(p).lower(),
    )
    if not files:
        return []

    selection = _load_image_selections(export_root).get(gs9, {})
    main_index = selection.get("main")
    if not isinstance(main_index, int):
        main_index = 1 if len(files) > 1 else 0
    if main_index < 0 or main_index >= len(files):
        main_index = 1 if len(files) > 1 else 0

    ordered = [files[main_index]]
    seen = {main_index}

    additional = selection.get("additional") if isinstance(selection.get("additional"), list) else []
    for idx in additional:
        if isinstance(idx, int) and 0 <= idx < len(files) and idx not in seen:
            ordered.append(files[idx])
            seen.add(idx)

    for idx, file_path in enumerate(files):
        if idx in seen:
            continue
        ordered.append(file_path)

    return ordered[:10]


def _resolve_public_image_url(image_path: str) -> str:
    cached = _COUPANG_PUBLIC_IMAGE_CACHE.get(image_path)
    if cached:
        return cached

    from app.services.naver_commerce import upload_image_url

    public_url = upload_image_url(image_path)
    _COUPANG_PUBLIC_IMAGE_CACHE[image_path] = public_url
    return public_url


def _build_fallback_image_urls(row: dict, detail_html: str) -> list[str]:
    import re as _re

    product_img_urls = []
    representative_candidates = []
    for value in (row.get("이미지등록(목록)"), row.get("이미지등록(상세)")):
        if not value:
            continue
        representative_candidates.extend(
            [u.strip() for u in _re.split(r"[|\n]", str(value)) if u and str(u).strip()]
        )

    detail_image_candidates = []
    if detail_html:
        html_imgs = _re.findall(r"<img[^>]+src=[\"']([^\"']+)", detail_html)
        gs_code = row.get("자체 상품코드") or ""
        for u in html_imgs:
            if gs_code and gs_code in u and _re.search(r"/\d+\.(jpg|jpeg|png|bmp|webp)$", u, _re.IGNORECASE):
                detail_image_candidates.append(u)

    seen_candidates = set()
    for candidate in representative_candidates + detail_image_candidates:
        candidate = str(candidate).strip()
        if not candidate or candidate in seen_candidates:
            continue
        seen_candidates.add(candidate)
        product_img_urls.append(candidate)

    return product_img_urls[:10]

def _build_coupang_image_urls(row: dict, detail_html: str) -> list[str]:
    local_images = _pick_local_listing_images(row)
    if local_images:
        public_urls = []
        first_error = None
        for image_path in local_images[:10]:
            try:
                public_urls.append(_resolve_public_image_url(image_path))
            except Exception as ex:
                if first_error is None:
                    first_error = ex
        if public_urls:
            return public_urls[:10]
        if first_error:
            raise ValueError(f"쿠팡 이미지 공개 URL 준비 실패: {first_error}") from first_error
        raise ValueError("쿠팡 이미지 공개 URL 준비 실패")

    return _build_fallback_image_urls(row, detail_html)


def read_source_file(file_path: str) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    headers = {c.column: c.value for c in ws[1] if c.value}
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for col, name in headers.items():
            row[name] = ws.cell(row=r, column=col).value
        row["_row_num"] = r
        row["_source_file_path"] = file_path
        row["_export_root"] = _resolve_export_root(file_path)
        rows.append(row)
    return rows

def parse_options(option_str: str | None, extra_price_str: str | None = None) -> list[dict]:
    if not option_str:
        return []
    m = re.findall(r"([A-Z])\s+([^,}]+)", option_str)
    prices = []
    if extra_price_str:
        # 구분자: , 또는 |
        for p in re.split(r"[,|]", str(extra_price_str)):
            p = p.strip()
            if p:
                try:
                    prices.append(int(float(p)))
                except ValueError:
                    prices.append(0)
    options = []
    for i, (_label, value) in enumerate(m):
        price = prices[i] if i < len(prices) else 0
        options.append({"name": value.strip(), "price": price})
    return options


def build_notice_content(category_meta: dict) -> list[dict]:
    notices = category_meta.get("data", {}).get("noticeCategories", [])
    if not notices:
        return []
    notice = notices[0]
    notice_name = notice.get("noticeCategoryName", "")
    details = notice.get("noticeCategoryDetailNames", [])
    return [
        {"noticeCategoryName": notice_name, "noticeCategoryDetailName": d.get("noticeCategoryDetailName", ""), "content": "상세페이지 참조"}
        for d in details
    ]


def build_attributes(category_meta: dict) -> list[dict]:
    attrs = category_meta.get("data", {}).get("attributes", [])
    result = []
    for a in attrs:
        if a["required"] != "MANDATORY":
            continue
        if a.get("exposed") == "EXPOSED":
            continue
        attr_name = a["attributeTypeName"]
        if a["inputType"] == "SELECT" and a.get("inputValues"):
            values = a["inputValues"]
            val = values[0].get("inputValueName", str(values[0])) if isinstance(values[0], dict) else str(values[0])
        elif attr_name in ("수량", "총 수량"):
            val = "1"
        elif attr_name == "색상":
            val = "기타"
        elif a["dataType"] == "NUMBER":
            val = "1"
        else:
            val = "상세페이지 참조"

        entry = {"attributeTypeName": attr_name, "attributeValueName": val}
        if a.get("basicUnit") and a["basicUnit"] != "없음":
            entry["unitCodeName"] = a["basicUnit"]
        result.append(entry)
    return result


def build_coupang_product(row: dict, category_code: int, category_meta: dict) -> dict:
    # 상품명: 키워드 가공된 이름 우선 사용
    product_name = (
        row.get("상품명")
        or row.get("최종키워드2차")
        or row.get("1차키워드")
        or ""
    )
    product_name = str(product_name).strip()
    display_name = product_name[:100]
    sale_price = max(int(row.get("판매가", 0) or 0), 1000)
    original_price = int(row.get("소비자가", 0) or 0)
    if original_price < sale_price:
        original_price = sale_price

    # 이미지: 로컬 가공본이 있으면 public URL로 변환해서 우선 사용
    detail_html = row.get("상품 상세설명") or row.get("상세설명") or ""
    product_img_urls = _build_coupang_image_urls(row, detail_html)

    images = []
    for idx, u in enumerate(product_img_urls[:10]):
        img_type = "REPRESENTATION" if idx == 0 else "DETAIL"
        images.append({"imageOrder": idx, "imageType": img_type, "vendorPath": u})

    options = parse_options(row.get("옵션입력"), row.get("옵션추가금"))
    notice_content = build_notice_content(category_meta)
    attributes = build_attributes(category_meta)

    search_tags = row.get("쿠팡검색태그") or row.get("검색어설정") or ""
    tag_list = [t.strip() for t in search_tags.replace(",", " ").split() if t.strip()][:10]

    mgmt_name = display_name
    detail_html = row.get("상품 상세설명") or row.get("상세설명") or ""
    ext_sku = row.get("자체 상품코드") or ""

    def _make_item(item_name: str, s_price: int, o_price: int, sku: str) -> dict:
        return {
            "itemName": item_name,
            "originalPrice": o_price,
            "salePrice": s_price,
            "maximumBuyCount": 9999,
            "maximumBuyForPerson": 9999,
            "outboundShippingTimeDay": 2,
            "maximumBuyForPersonPeriod": 1,
            "unitCount": 1,
            "adultOnly": "EVERYONE",
            "taxType": "TAX",
            "parallelImported": "NOT_PARALLEL_IMPORTED",
            "overseasPurchased": "NOT_OVERSEAS_PURCHASED",
            "pccNeeded": False,
            "externalVendorSku": sku,
            "barcode": "",
            "emptyBarcode": True,
            "emptyBarcodeReason": "",
            "notices": notice_content,
            "attributes": attributes,
            "contents": [{"contentsType": "HTML", "contentDetails": [{"content": detail_html, "detailType": "TEXT"}]}],
            "images": images,
            "searchTags": tag_list,
        }

    items = []
    if options:
        for i, opt in enumerate(options):
            items.append(_make_item(opt["name"], sale_price + opt["price"], original_price + opt["price"], f"{ext_sku}_{i+1}" if ext_sku else ""))
    else:
        items.append(_make_item(display_name, sale_price, original_price, ext_sku))

    return {
        "displayCategoryCode": category_code,
        "sellerProductName": mgmt_name,
        "vendorId": VENDOR_ID,
        "saleStartedAt": "2020-01-01T00:00:00",
        "saleEndedAt": "2099-12-31T00:00:00",
        "displayProductName": display_name,
        "brand": "자체브랜드",
        "generalProductName": "",
        "productGroup": "",
        "deliveryMethod": "SEQUENCIAL",
        "deliveryCompanyCode": "CJGLS",
        "deliveryChargeType": "FREE",
        "deliveryCharge": 0,
        "freeShipOverAmount": 0,
        "deliveryChargeOnReturn": 3000,
        "returnCharge": 3000,
        "outboundShippingPlaceCode": OUTBOUND_CODE,
        "returnCenterCode": RETURN_CENTER_CODE,
        "returnChargeName": "명일우진반품",
        "companyContactNumber": "010-2324-8352",
        "returnZipCode": "05287",
        "returnAddress": "서울특별시 강동구 상일로 74",
        "returnAddressDetail": "고덕리엔파크3단지아파트 고덕리엔파크 321동 CJ대한통운 명일우진대리점",
        "remoteAreaDeliverable": "Y",
        "unionDeliveryType": "UNION_DELIVERY",
        "vendorUserId": "rkghrud",
        "afterServiceInformation": "010-2324-8352",
        "afterServiceContactNumber": "010-2324-8352",
        "requested": True,
        "items": items,
        "requiredDocuments": [],
        "extraInfoMessage": "",
        "manufacture": "",
    }


# ─── 업로드 파이프라인 (GUI 콜백 버전) ─────────

@dataclass
class CoupangUploadConfig:
    file_path: str = ""
    row_start: int = 0          # 0이면 전체
    row_end: int = 0
    dry_run: bool = True

@dataclass
class CoupangUploadResult:
    row: int = 0
    name: str = ""
    status: str = ""
    category: str = ""
    seller_product_id: str = ""
    error: str = ""


def run_coupang_upload(
    config: CoupangUploadConfig,
    status_cb: Callable[[str], None] | None = None,
    progress_cb: Callable[[int], None] | None = None,
) -> list[CoupangUploadResult]:
    """GUI에서 호출하는 쿠팡 업로드 메인 함수"""

    def _status(msg: str) -> None:
        if status_cb:
            status_cb(msg)

    def _progress(pct: int) -> None:
        if progress_cb:
            progress_cb(pct)

    _status("가공파일 읽는 중...")
    rows = read_source_file(config.file_path)
    total = len(rows)
    _status(f"{total}개 상품 로드 완료")

    # 행 필터
    if config.row_start > 0:
        end = config.row_end if config.row_end > 0 else config.row_start
        target_rows = [r for r in rows if config.row_start <= (r["_row_num"] - 1) <= end]
    else:
        target_rows = rows

    target_count = len(target_rows)
    _status(f"처리 대상: {target_count}개")
    results: list[CoupangUploadResult] = []
    category_cache: dict[int, dict] = {}

    # 카테고리 추천 (배치 8건 + 1초 대기, 쿠팡 API 초당 10건 제한)
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import threading

    # 엑셀에 쿠팡카테고리코드가 있는 행은 API 호출 스킵
    rows_need_predict = []
    cat_results_ordered: list[tuple[dict, dict]] = [({}, {})] * target_count
    for i, row in enumerate(target_rows):
        preset_cat = row.get("쿠팡카테고리코드") or row.get("쿠팡카테고리")
        if preset_cat:
            cat_code = int(float(str(preset_cat)))
            cat_results_ordered[i] = (row, {"data": {"predictedCategoryId": cat_code, "predictedCategoryName": f"엑셀지정({cat_code})", "autoCategorizationPredictionResultType": "SUCCESS"}, "_preset": True})
            _status(f"  행{row['_row_num']}: 엑셀 카테고리 사용 ({cat_code})")
        else:
            rows_need_predict.append((i, row))

    if rows_need_predict:
        _status(f"카테고리 추천 중... ({len(rows_need_predict)}건 API 호출)")
        cat_lock = threading.Lock()

        BATCH_SIZE = 8
        for batch_start in range(0, len(rows_need_predict), BATCH_SIZE):
            batch_end = min(batch_start + BATCH_SIZE, len(rows_need_predict))
            batch = rows_need_predict[batch_start:batch_end]
            batch_t0 = time.time()

            def _predict_one(idx: int, row: dict) -> tuple[int, dict, dict]:
                product_name = row.get("상품명", "")
                cat_result = predict_category(product_name)
                return idx, row, cat_result

            with ThreadPoolExecutor(max_workers=BATCH_SIZE) as pool:
                futures = [pool.submit(_predict_one, i, row) for i, row in batch]
                for fut in as_completed(futures):
                    idx, row, cat_result = fut.result()
                    cat_results_ordered[idx] = (row, cat_result)

        _status(f"[{batch_end}/{target_count}] 카테고리 추천 중...")
        _progress(int(batch_end * 40 / target_count))

        # 초당 10건 제한: 배치 처리 후 최소 1초 대기
        elapsed = time.time() - batch_t0
        if elapsed < 1.0:
            time.sleep(1.0 - elapsed)

    # 결과 처리 (순서대로)
    for i in range(target_count):
        row, cat_result = cat_results_ordered[i]
        product_name = row.get("상품명", "")
        short_name = product_name[:50]
        row_num = row["_row_num"]

        if "_error" in cat_result:
            results.append(CoupangUploadResult(row=row_num, name=short_name, status="CATEGORY_FAIL", error=cat_result.get("_msg", "")[:200]))
            continue

        data = cat_result.get("data", {})
        cat_code = data.get("predictedCategoryId")
        cat_name = data.get("predictedCategoryName", "")
        result_type = data.get("autoCategorizationPredictionResultType", "")

        if not cat_code or result_type != "SUCCESS":
            results.append(CoupangUploadResult(row=row_num, name=short_name, status="CATEGORY_UNCERTAIN", category=cat_name))
            continue

        if cat_code not in category_cache:
            meta = get_category_meta(cat_code)
            category_cache[cat_code] = meta if "_error" not in meta else {"data": {"attributes": [], "noticeCategories": []}}

        row["_category_code"] = cat_code
        row["_category_name"] = cat_name
        row["_category_meta"] = category_cache[cat_code]

    _status(f"카테고리 추천 완료")

    # JSON 생성
    _status("상품 JSON 생성 중...")
    products = []
    for row in target_rows:
        if "_category_code" not in row:
            continue
        product_json = build_coupang_product(row, row["_category_code"], row["_category_meta"])
        products.append({"row": row["_row_num"], "name": row.get("상품명", "")[:50], "sku": row.get("자체 상품코드", ""), "category": f"[{row['_category_code']}] {row['_category_name']}", "json": product_json})
    _progress(50)
    _status(f"JSON 생성 완료: {len(products)}개")

    # 등록 (배치 5건 + 1초 대기)
    if not config.dry_run:
        prod_count = len(products)
        _status(f"쿠팡 등록 시작 ({prod_count}개)...")
        REG_BATCH = 5
        reg_results_ordered: list[CoupangUploadResult] = [CoupangUploadResult()] * prod_count

        def _register_one(idx: int, p: dict) -> tuple[int, CoupangUploadResult]:
            resp = create_product(p["json"])
            if "_error" in resp:
                return idx, CoupangUploadResult(row=p["row"], name=p["name"], status="REGISTER_FAIL", error=resp["_msg"][:200])
            code = resp.get("code", "")
            if code == "SUCCESS":
                spid = str(resp.get("data", ""))
                return idx, CoupangUploadResult(row=p["row"], name=p["name"], status="SUCCESS", seller_product_id=spid)
            return idx, CoupangUploadResult(row=p["row"], name=p["name"], status=f"FAIL_{code}", error=resp.get("message", "")[:200])

        for batch_start in range(0, prod_count, REG_BATCH):
            batch_end = min(batch_start + REG_BATCH, prod_count)
            batch_t0 = time.time()
            with ThreadPoolExecutor(max_workers=REG_BATCH) as pool:
                futures = [pool.submit(_register_one, i, products[i]) for i in range(batch_start, batch_end)]
                for fut in as_completed(futures):
                    idx, r = fut.result()
                    reg_results_ordered[idx] = r
            _status(f"[{batch_end}/{prod_count}] 등록 중...")
            _progress(50 + int(batch_end * 50 / prod_count))
            elapsed = time.time() - batch_t0
            if elapsed < 1.0:
                time.sleep(1.0 - elapsed)

        for i in range(prod_count):
            r = reg_results_ordered[i]
            results.append(r)
            _status(f"[{i+1}/{prod_count}] {r.status} - {r.name}")
    else:
        _status("DRY RUN 완료 - 등록하지 않음")
        for p in products:
            results.append(CoupangUploadResult(row=p["row"], name=p["name"], status="DRY_RUN", category=p["category"]))
        _progress(100)

    _progress(100)
    return results
